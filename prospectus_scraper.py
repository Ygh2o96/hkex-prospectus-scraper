#!/usr/bin/env python3
"""
HKEX Listed Company Prospectus Scraper
=======================================
Uses Playwright to automate the HKEX Title Search, querying by stock code
+ "Listing Documents" category to find prospectus PDFs.

Data flow:
  1. Download NLR Excel for year(s) → stock codes + prospectus dates
  2. Load activestock JSON → map stock codes to HKEX internal stockIds
  3. Playwright → title search per stock → extract PDF links
  4. Download PDFs

Designed for GitHub Actions with Playwright pre-installed.
"""

import argparse
import hashlib
import json
import logging
import os
import re
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

import requests

try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages -q")
    import openpyxl

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    os.system(f"{sys.executable} -m pip install playwright --break-system-packages -q")
    os.system(f"{sys.executable} -m playwright install chromium")
    from playwright.sync_api import sync_playwright

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

BASE_URL = "https://www1.hkexnews.hk"
NLR_BASE = "https://www2.hkexnews.hk/-/media/HKEXnews/Homepage/New-Listings/New-Listing-Information/New-Listing-Report/Main"
ACTIVESTOCK_URL = f"{BASE_URL}/ncms/script/eds/activestock_sehk_e.json"
TITLESEARCH_URL = f"{BASE_URL}/search/titlesearch.xhtml"
LISTING_DOCS_CATEGORY = "30000"

# HKEX Title Search tier2 subcategory codes (under tier1=30000 Listing Documents)
# Source: /ncms/script/eds/tiertwo_e.json
TIER2_CODES = {
    "prospectus": "30700",   # Offer for Subscription = Global Offering / Prospectus
    "offer-sale": "30600",   # Offer for Sale
    "introduction": "30500", # Introduction
    "rights": "31100",       # Rights Issue
    "open-offer": "30800",   # Open Offer
    "supplemental": "31200", # Supplementary Listing Document
    "all": "-2",             # All subcategories
}

DOWNLOAD_DIR = Path(os.environ.get("DOWNLOAD_DIR", "downloads/prospectuses"))
STATE_FILE = Path(os.environ.get("STATE_FILE", "state/prospectus_downloaded.json"))
RATE_LIMIT = float(os.environ.get("RATE_LIMIT_SECONDS", "2.0"))
DRY_RUN = os.environ.get("DRY_RUN", "false").lower() == "true"

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s",
                    datefmt="%Y-%m-%d %H:%M:%S")
log = logging.getLogger(__name__)

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer": "https://www1.hkexnews.hk/search/titlesearch.xhtml",
})

# ---------------------------------------------------------------------------
# State
# ---------------------------------------------------------------------------

def load_state() -> dict:
    if STATE_FILE.exists():
        return json.loads(STATE_FILE.read_text())
    return {"downloaded": {}, "searched": {}, "last_run": None}

def save_state(state: dict):
    STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    state["last_run"] = datetime.utcnow().isoformat()
    STATE_FILE.write_text(json.dumps(state, indent=2, ensure_ascii=False))

# ---------------------------------------------------------------------------
# NLR Excel
# ---------------------------------------------------------------------------

def get_nlr_url(year: int) -> str:
    if year >= 2020:
        return f"{NLR_BASE}/NLR{year}_Eng.xlsx"
    elif year >= 2005:
        return f"{NLR_BASE}/NLR{year}_Eng.xls"
    else:
        return f"{NLR_BASE}/{year}.xls"

def download_nlr(year: int, force: bool = False) -> Path:
    url = get_nlr_url(year)
    ext = ".xlsx" if url.endswith(".xlsx") else ".xls"
    local = Path(f"/tmp/NLR{year}{ext}")
    if local.exists() and not force:
        return local
    log.info(f"Downloading NLR {year}: {url}")
    resp = SESSION.get(url, timeout=30)
    resp.raise_for_status()
    local.write_bytes(resp.content)
    return local

def parse_nlr(filepath: Path) -> list[dict]:
    """Parse NLR Excel → [{stock_code, company, prospectus_date, listing_date}]."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active or wb[wb.sheetnames[0]]
    listings = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        stock_code, company, prosp_date, list_date = row[1], row[2], row[3], row[4]
        if not stock_code or not company:
            continue
        if str(stock_code).strip() in ('"', ''):
            continue
        code = str(stock_code).strip().zfill(5)
        if not re.match(r'^\d{5}$', code):
            continue

        def to_date(val):
            if hasattr(val, 'strftime'):
                return val.strftime('%Y-%m-%d')
            if isinstance(val, str):
                for fmt in ('%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d'):
                    try:
                        return datetime.strptime(val.strip(), fmt).strftime('%Y-%m-%d')
                    except ValueError:
                        continue
            return None

        listings.append({
            "stock_code": code,
            "company": str(company).strip(),
            "prospectus_date": to_date(prosp_date),
            "listing_date": to_date(list_date),
        })
    wb.close()
    # Deduplicate (NLR has merged cells that create duplicate rows)
    seen = set()
    deduped = []
    for l in listings:
        key = l["stock_code"]
        if key not in seen:
            seen.add(key)
            deduped.append(l)
    log.info(f"  Parsed {len(deduped)} listings from {filepath.name}")
    return deduped

# ---------------------------------------------------------------------------
# Stock code → HKEX internal ID
# ---------------------------------------------------------------------------

_stock_map = None

def load_stock_map() -> dict:
    global _stock_map
    if _stock_map:
        return _stock_map
    log.info("Loading activestock mapping...")
    _stock_map = {}
    for endpoint in [ACTIVESTOCK_URL,
                     ACTIVESTOCK_URL.replace("activestock", "inactivestock")]:
        try:
            data = SESSION.get(endpoint, timeout=30).json()
            for item in data:
                code = str(item.get("c", "")).strip()
                iid = item.get("i")
                if code and iid:
                    _stock_map.setdefault(code, iid)
        except Exception:
            pass
    log.info(f"  Loaded {len(_stock_map)} stock mappings")
    return _stock_map

# ---------------------------------------------------------------------------
# Playwright search
# ---------------------------------------------------------------------------

def search_listing_docs(browser, stock_code: str, internal_id: int) -> list[dict]:
    """
    Automate HKEX Title Search: stock + category=Listing Documents.
    Returns list of {title, url, date, headline} where headline is the
    subcategory text (e.g. "Offer for Subscription", "Rights Issue").
    """
    page = browser.new_page()
    page.set_default_timeout(60000)
    results = []

    try:
        url = (f"{TITLESEARCH_URL}?lang=EN&market=SEHK"
               f"&stockId={internal_id}&category={LISTING_DOCS_CATEGORY}")
        page.goto(url, wait_until="domcontentloaded", timeout=60000)
        page.wait_for_timeout(4000)

        # Click search
        for sel in ["a.filter__btn-apply:not(.btn-disable)",
                    ".filter__btn-apply:not(.btn-disable)",
                    "a[class*='btn-apply']"]:
            btn = page.query_selector(sel)
            if btn and btn.is_visible():
                btn.click()
                break

        try:
            page.wait_for_selector(
                "#titleSearchResultPanel .news_headline,"
                "#titleSearchResultPanel a[href*='.pdf'],"
                ".no-match-text", timeout=20000)
        except Exception:
            pass
        page.wait_for_timeout(2000)

        # Extract PDFs with headline subcategory text
        results = page.evaluate("""() => {
            const links = [];
            // Each result row has: date, stock info, headline category, document link
            const panel = document.getElementById('titleSearchResultPanel');
            if (!panel) return links;

            // Get all PDF links
            panel.querySelectorAll('a[href*=".pdf"]').forEach(a => {
                let href = a.getAttribute('href') || '';
                if (href.startsWith('/')) href = 'https://www1.hkexnews.hk' + href;

                // Walk up to find the result container (row/div)
                let container = a.closest('tr') || a.closest('.row') || a.parentElement?.parentElement;
                const containerText = container ? container.innerText : '';

                // Extract headline: "Listing Documents - [Offer for Subscription]"
                const hlMatch = containerText.match(/Listing Documents\\s*-\\s*\\[([^\\]]+)\\]/i);
                const headline = hlMatch ? hlMatch[1].trim() : '';

                // Extract date: DD/MM/YYYY
                const dtMatch = containerText.match(/(\\d{2}\\/\\d{2}\\/\\d{4})/);
                const date = dtMatch ? dtMatch[1] : '';

                links.push({
                    url: href,
                    title: a.innerText.trim().substring(0, 120),
                    date: date,
                    headline: headline
                });
            });
            return links;
        }""")

        log.info(f"  {stock_code}: {len(results)} result(s) from HKEX")

    except Exception as e:
        log.error(f"  {stock_code}: search failed — {e}")
    finally:
        page.close()

    return results

# ---------------------------------------------------------------------------
# Download
# ---------------------------------------------------------------------------

def download_pdf(url: str, dest: Path, state: dict) -> bool:
    url_hash = hashlib.md5(url.encode()).hexdigest()
    if url_hash in state["downloaded"]:
        return False
    if dest.exists():
        state["downloaded"][url_hash] = {"url": url, "path": str(dest),
                                          "ts": datetime.utcnow().isoformat()}
        return False
    if DRY_RUN:
        log.info(f"    [DRY RUN] {dest.name}")
        return True

    try:
        time.sleep(RATE_LIMIT)
        resp = SESSION.get(url, timeout=120, stream=True)
        resp.raise_for_status()
        dest.parent.mkdir(parents=True, exist_ok=True)
        with open(dest, "wb") as f:
            for chunk in resp.iter_content(65536):
                f.write(chunk)
        size_mb = dest.stat().st_size / (1024 * 1024)
        log.info(f"    ✓ {dest.name} ({size_mb:.1f} MB)")
        state["downloaded"][url_hash] = {"url": url, "path": str(dest),
                                          "ts": datetime.utcnow().isoformat()}
        return True
    except Exception as e:
        log.error(f"    ✗ {dest.name} — {e}")
        return False

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Cron management
# ---------------------------------------------------------------------------

CRON_TAG = "# hkex-prospectus-scraper"
LAUNCHD_LABEL = "com.hkex.prospectus-scraper"

def _manage_cron(args):
    """Install or remove a daily scheduled job. Uses launchd on macOS, cron on Linux."""
    import subprocess, shutil, platform

    is_mac = platform.system() == "Darwin"

    # Build the command from current args
    script = Path(__file__).resolve()
    python = shutil.which("python3") or shutil.which("python") or sys.executable

    cmd_parts = [python, str(script)]
    if args.latest:
        cmd_parts += ["--latest", str(args.latest)]
    else:
        cmd_parts += ["--latest", "30"]

    cmd_parts += ["--filter", args.filter]
    if args.filter == "smart":
        cmd_parts += ["--min-size-mb", str(args.min_size_mb)]
    elif args.filter == "top-n":
        cmd_parts += ["--top-n", str(args.top_n)]

    cmd_parts += ["--doc-type", args.doc_type]

    output_dir = args.output_dir or str(DOWNLOAD_DIR)
    cmd_parts += ["--output-dir", output_dir]

    try:
        hh, mm = args.cron_time.split(":")
        cron_hour, cron_min = int(hh), int(mm)
    except Exception:
        cron_hour, cron_min = 11, 0

    if args.remove_cron:
        if is_mac:
            _launchd_remove()
        else:
            _cron_remove()
        return

    if is_mac:
        _launchd_install(cmd_parts, output_dir, cron_hour, cron_min, args.cron_time)
    else:
        _cron_install(cmd_parts, output_dir, cron_hour, cron_min, args.cron_time)


def _launchd_install(cmd_parts, output_dir, hour, minute, time_str):
    """Install a macOS launchd plist — survives sleep, fires on wake if missed."""
    import plistlib, subprocess

    out_dir = Path(output_dir).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    log_path = out_dir / "cron.log"

    plist = {
        "Label": LAUNCHD_LABEL,
        "ProgramArguments": cmd_parts,
        "StartCalendarInterval": {"Hour": hour, "Minute": minute},
        "StandardOutPath": str(log_path),
        "StandardErrorPath": str(log_path),
        "WorkingDirectory": str(Path(cmd_parts[1]).parent),
        "EnvironmentVariables": {
            "PATH": "/usr/local/bin:/usr/bin:/bin:/opt/homebrew/bin",
        },
    }

    plist_dir = Path.home() / "Library" / "LaunchAgents"
    plist_dir.mkdir(parents=True, exist_ok=True)
    plist_path = plist_dir / f"{LAUNCHD_LABEL}.plist"

    # Unload old version if exists
    if plist_path.exists():
        subprocess.run(["launchctl", "unload", str(plist_path)],
                       capture_output=True)

    with open(plist_path, "wb") as f:
        plistlib.dump(plist, f)

    result = subprocess.run(["launchctl", "load", str(plist_path)],
                            capture_output=True, text=True)

    if result.returncode == 0:
        log.info(f"macOS launchd job installed (daily at {time_str}):")
        log.info(f"  Plist  → {plist_path}")
        log.info(f"  Output → {out_dir}")
        log.info(f"  Logs   → {log_path}")
        log.info(f"\n  ✓ Will fire on wake if laptop was asleep at {time_str}")
        log.info(f"\n  Test now:  launchctl start {LAUNCHD_LABEL}")
        log.info(f"  Remove:    python {Path(cmd_parts[1]).name} --remove-cron")
    else:
        log.error(f"launchctl load failed: {result.stderr}")


def _launchd_remove():
    import subprocess
    plist_path = Path.home() / "Library" / "LaunchAgents" / f"{LAUNCHD_LABEL}.plist"
    if plist_path.exists():
        subprocess.run(["launchctl", "unload", str(plist_path)], capture_output=True)
        plist_path.unlink()
        log.info(f"launchd job removed: {plist_path}")
    else:
        log.info("No launchd job found.")


def _cron_install(cmd_parts, output_dir, hour, minute, time_str):
    """Install a Linux cron job."""
    import subprocess
    cron_cmd = " ".join(cmd_parts)
    cron_line = f'{minute} {hour} * * * {cron_cmd} >> {Path(output_dir).expanduser()}/cron.log 2>&1 {CRON_TAG}'

    result = subprocess.run(["crontab", "-l"], capture_output=True, text=True)
    existing = result.stdout if result.returncode == 0 else ""
    lines = [l for l in existing.splitlines() if CRON_TAG not in l]
    lines.append(cron_line)

    proc = subprocess.run(["crontab", "-"], input="\n".join(lines) + "\n",
                          capture_output=True, text=True)
    if proc.returncode == 0:
        log.info(f"Cron job installed (daily at {time_str}):")
        log.info(f"  {cron_line}")
        log.info(f"\n  ⚠ Cron does NOT fire if machine is asleep at {time_str}")
        log.info(f"  Remove: python {Path(cmd_parts[1]).name} --remove-cron")
    else:
        log.error(f"crontab failed: {proc.stderr}")


def _cron_remove():
    import subprocess
    result = subprocess.run(["crontab", "-l"], capture_output=True, text=True)
    if result.returncode != 0:
        log.info("No crontab found.")
        return
    lines = [l for l in result.stdout.splitlines() if CRON_TAG not in l]
    subprocess.run(["crontab", "-"], input="\n".join(lines) + "\n",
                   capture_output=True, text=True)
    log.info("Cron job removed.")


# ---------------------------------------------------------------------------
# Date parsing
# ---------------------------------------------------------------------------

def parse_flexible_date(s: str) -> datetime:
    """Parse dates in various formats: 01Jan2024, 01/01/2024, 2024-01-01, 01-Jan-2024."""
    s = s.strip()
    for fmt in ("%d%b%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d%B%Y", "%Y%m%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    raise ValueError(f"Cannot parse date: '{s}'. Use formats like 01Jan2024, 01/01/2024, 2024-01-01")


def main():
    parser = argparse.ArgumentParser(
        description="HKEX Prospectus Scraper (Playwright)",
        epilog="""Examples:
  # Top-3 for all of 2026
  %(prog)s --years 2026 --filter top-n

  # Date range
  %(prog)s --date-range 01Jan2025-31Mar2026 --filter smart

  # Latest 30 days (for cron)
  %(prog)s --latest 30 --filter top-n --top-n 3

  # Install daily cron job
  %(prog)s --install-cron --latest 30 --filter top-n --output-dir ~/HKEX_Prospectuses
""",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    # Source selection (at least one required)
    src = parser.add_argument_group("Source selection (pick one or combine)")
    src.add_argument("--years", nargs="+", type=int, default=None,
                     help="Year(s) to scrape, e.g. --years 2024 2025 2026")
    src.add_argument("--date-range", type=str, default=None,
                     help="Listing date range, e.g. 01Jan2025-31Dec2025 or 01/01/2025-31/12/2025")
    src.add_argument("--latest", type=int, default=None,
                     help="Only process listings from the last N days (ideal for cron)")
    parser.add_argument("--stock", type=str, default=None,
                        help="Single stock code to scrape (testing)")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--date-margin", type=int, default=7,
                        help="Days around prospectus date to search HKEX (default: 7)")
    parser.add_argument("--filter", type=str, default="smart",
                        choices=["all", "smart", "top-n"],
                        help="Download filter: all=everything, smart=skip sub-MB files, "
                             "top-n=largest N files only (default: smart)")
    parser.add_argument("--min-size-mb", type=float, default=1.0,
                        help="Min file size in MB for 'smart' filter (default: 1.0)")
    parser.add_argument("--top-n", type=int, default=3,
                        help="Number of largest files to keep for 'top-n' filter (default: 3)")
    parser.add_argument("--doc-type", type=str, default="prospectus",
                        choices=["prospectus", "all-listed", "ap-phip", "everything"],
                        help="Which documents to download: "
                             "prospectus=only Global Offering doc (fastest, server-side filter), "
                             "all-listed=all listed-co Listing Documents, "
                             "ap-phip=only AP/PHIP (sehk*/gem* files), "
                             "everything=no filter (default: prospectus)")
    parser.add_argument("--output-dir", type=str, default=None,
                        help="Override download directory (default: downloads/prospectuses)")

    # Cron management
    parser.add_argument("--install-cron", action="store_true",
                        help="Install a daily cron job and exit.")
    parser.add_argument("--cron-time", type=str, default="11:00",
                        help="Time for cron job in HH:MM (default: 11:00)")
    parser.add_argument("--remove-cron", action="store_true",
                        help="Remove the installed cron job and exit")
    args = parser.parse_args()

    # ---- Cron management ----
    if args.install_cron or args.remove_cron:
        _manage_cron(args)
        return

    # ---- Override output dir ----
    global DOWNLOAD_DIR
    if args.output_dir:
        base = Path(args.output_dir).expanduser().resolve()
        # Auto-append 'prospectuses' if not already in the path
        if base.name != "prospectuses":
            DOWNLOAD_DIR = base / "prospectuses"
        else:
            DOWNLOAD_DIR = base

    if args.dry_run:
        global DRY_RUN
        DRY_RUN = True

    # ---- Determine which NLR years to fetch ----
    date_filter_from = None
    date_filter_to = None

    if args.date_range:
        parts = args.date_range.split("-", 1)
        if len(parts) != 2:
            # Try other separators
            for sep in (" to ", "~", ".."):
                if sep in args.date_range:
                    parts = args.date_range.split(sep, 1)
                    break
        if len(parts) != 2:
            log.error("Invalid --date-range format. Use: 01Jan2025-31Dec2025")
            return
        date_filter_from = parse_flexible_date(parts[0])
        date_filter_to = parse_flexible_date(parts[1])
        log.info(f"Date range filter: {date_filter_from.strftime('%Y-%m-%d')} → {date_filter_to.strftime('%Y-%m-%d')}")

    if args.latest:
        date_filter_to = datetime.now()
        date_filter_from = date_filter_to - timedelta(days=args.latest)
        log.info(f"Latest {args.latest} days: {date_filter_from.strftime('%Y-%m-%d')} → {date_filter_to.strftime('%Y-%m-%d')}")

    # Determine years to fetch NLR files for
    if args.years:
        years = args.years
    elif date_filter_from and date_filter_to:
        years = list(range(date_filter_from.year, date_filter_to.year + 1))
    else:
        # Default: current year
        years = [datetime.now().year]
        log.info(f"No --years/--date-range/--latest specified, defaulting to {years[0]}")

    state = load_state()

    # 1. Parse NLR files
    all_listings = []
    force_fresh = args.latest is not None  # always re-download NLR in cron mode
    for year in years:
        try:
            nlr_path = download_nlr(year, force=force_fresh)
            all_listings.extend(parse_nlr(nlr_path))
        except Exception as e:
            log.error(f"NLR {year} failed: {e}")

    if not all_listings:
        log.error("No listings found.")
        return

    # Apply date range filter on listing_date
    if date_filter_from and date_filter_to:
        before = len(all_listings)
        filtered = []
        for l in all_listings:
            ld = l.get("listing_date")
            if not ld:
                continue
            try:
                ld_dt = datetime.strptime(ld, "%Y-%m-%d")
                if date_filter_from <= ld_dt <= date_filter_to:
                    filtered.append(l)
            except ValueError:
                continue
        all_listings = filtered
        log.info(f"Date filter: {before} → {len(all_listings)} listings")

    if args.stock:
        code = args.stock.zfill(5)
        all_listings = [l for l in all_listings if l["stock_code"] == code]

    log.info(f"\nListings to process: {len(all_listings)}")

    # 2. Load stock map
    stock_map = load_stock_map()

    # 3. Search + download
    success = skip = fail = 0

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu"]
        )

        for i, listing in enumerate(all_listings):
            code = listing["stock_code"]
            company = listing["company"]
            prosp_date = listing["prospectus_date"]
            search_key = f"{code}_{prosp_date}"

            if search_key in state.get("searched", {}) and not args.stock:
                skip += 1
                continue

            internal_id = stock_map.get(code)
            if not internal_id:
                log.warning(f"  {code} ({company[:30]}): no ID mapping")
                fail += 1
                continue

            log.info(f"\n[{i+1}/{len(all_listings)}] {code} {company[:40]} (prosp: {prosp_date})")

            # Search HKEX (returns all Listing Documents for this stock)
            docs = search_listing_docs(browser, code, internal_id)

            # Filter to PDFs only
            pdf_docs = [d for d in docs if d.get("url", "").endswith(".pdf")]

            # ── HEADLINE-BASED DOC TYPE FILTER (the real fix) ──
            # HKEX labels each result: "Offer for Subscription", "Rights Issue", etc.
            # Map --doc-type to headline keywords
            PROSPECTUS_HEADLINES = {"offer for subscription", "offer for sale",
                                    "placing of securities of a class new to listing"}

            before = len(pdf_docs)
            if args.doc_type == "prospectus":
                # Filter by headline text (primary) + filename fallback
                filtered = []
                for d in pdf_docs:
                    hl = d.get("headline", "").lower()
                    fname = d["url"].split("/")[-1].lower()
                    # Match by headline if available
                    if hl and hl in PROSPECTUS_HEADLINES:
                        filtered.append(d)
                    # If no headline extracted, use filename heuristic
                    elif not hl and not fname.startswith(("sehk", "gem")):
                        filtered.append(d)
                pdf_docs = filtered
            elif args.doc_type == "all-listed":
                pdf_docs = [d for d in pdf_docs
                            if not d["url"].split("/")[-1].lower().startswith(("sehk", "gem"))]
            elif args.doc_type == "ap-phip":
                pdf_docs = [d for d in pdf_docs
                            if d["url"].split("/")[-1].lower().startswith(("sehk", "gem"))]
            # "everything" = no filter

            if before != len(pdf_docs):
                log.info(f"  Filtered: {before} → {len(pdf_docs)} ({args.doc_type})")

            # ── Size filter (only when explicitly needed) ──
            if args.filter in ("smart", "top-n") and pdf_docs:
                log.info(f"  Checking sizes for {len(pdf_docs)} files...")
                for doc in pdf_docs:
                    try:
                        resp = SESSION.head(doc["url"], timeout=10, allow_redirects=True)
                        doc["size"] = int(resp.headers.get("Content-Length", 0))
                    except Exception:
                        doc["size"] = 0
                    time.sleep(0.15)
                if args.filter == "smart":
                    min_bytes = int(args.min_size_mb * 1024 * 1024)
                    pdf_docs = [d for d in pdf_docs if d.get("size", 0) >= min_bytes]
                elif args.filter == "top-n":
                    pdf_docs.sort(key=lambda d: d.get("size", 0), reverse=True)
                    pdf_docs = pdf_docs[:args.top_n]

            log.info(f"  → {len(pdf_docs)} file(s) to download")

            # Download
            safe_name = re.sub(r'[<>:"/\\|?*]', '_', company).strip()[:80]
            for doc in pdf_docs:
                url = doc["url"]
                filename = url.split("/")[-1]
                dest = DOWNLOAD_DIR / f"{code}_{safe_name}" / filename
                if download_pdf(url, dest, state):
                    success += 1

            state.setdefault("searched", {})[search_key] = {
                "docs_found": len(docs), "ts": datetime.utcnow().isoformat()
            }
            time.sleep(RATE_LIMIT)

        browser.close()

    save_state(state)

    log.info(f"\n{'='*60}")
    log.info(f"Processed: {len(all_listings)} | Downloaded: {success} | Skipped: {skip} | Failed: {fail}")

    summary_file = os.environ.get("GITHUB_STEP_SUMMARY")
    if summary_file:
        with open(summary_file, "a") as f:
            f.write(f"\n## Prospectus Scraper\n\n")
            f.write(f"- **Years**: {args.years}\n")
            f.write(f"- **Listings**: {len(all_listings)} | **Downloaded**: {success}\n")

if __name__ == "__main__":
    main()
